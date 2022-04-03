import openpyxl

def Getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)

def Getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)

def Readdata(file,sheetname,rowno,colmnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno,column=colmnno).value

def Writedata(file,sheetname,rowno,colmnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno,column=colmnno).value=data
    workbook.save(file)